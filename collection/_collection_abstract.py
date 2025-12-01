from openpyxl import Workbook
import os
from enum import Enum
from typing import List, Tuple
from abc import ABC, abstractmethod
from dataclasses import is_dataclass, fields

from typing import List, Dict, Optional, Type, TypeVar, Generic, Union, Callable, Any, Set, Tuple

from SANSPRO.model.model import Model, BlockAdapter
from object._object_abstract import Object

M = TypeVar('M', bound='Model')
T = TypeVar('T', bound='Object')
C = TypeVar('C', bound='Collection[Object]')

class Collection(Generic[T]):

    header: str = ""

    def __init__(self, objects: Optional[List[T]] = None):
        self.objects: List[T] = []
        self._index: Dict[int, T] = {}
        self._reverse_index: Dict[int, int] = {}

        if objects:
            self._initialize(objects)

    def _initialize(self, objects: List[T]):
        for obj in objects:
            self.add(obj)

    def add(self, obj: T):
        self.objects.append(obj)
        self._index[obj.index] = obj
        self._reverse_index[id(obj)] = obj.index

    def remove(self, obj: T):
        if obj in self.objects:
            self.objects.remove(obj)
            self._index.pop(obj.index, None)
            self._reverse_index.pop(id(obj), None)

    def extend(self, objs: List[T]):
        for obj in objs:
            self.add(obj)

    def index_list(self) -> List[int]:
        return sorted(self._index.keys())

    def get(self, index: int) -> Optional[T]:
        return self._index.get(index)

    def __getitem__(self, index: int) -> T:
        return self._index[index]

    def __contains__(self, index: int) -> bool:
        return index in self._index

    def __iter__(self):
        return iter(self.objects)
    

    # ==========================================================
    # SUMMARY UTILITIES
    # ==========================================================
    def summary(self) -> str:
        """Return a compact summary of all available indices."""
        ids = sorted(self._index.keys())
        return print(f"{self.header or 'Collection'}: {len(ids)} items, indices = {ids}")
    
    # ==========================================================
    # 🔹 Strict Elset Collector
    # ==========================================================

    def get_used_elsets(self) -> Set[int]:
        """
        Return all unique Elset indices referenced by this collection.
        Empty set if none found.
        """
        if not self.objects:
            return set()

        used: set[int] = set()

        for obj in self.objects:
            # Direct elset
            if hasattr(obj, "elset"):
                elset = getattr(obj, "elset")
                if elset is not None:
                    used.add(elset.index)
                continue

            # Nested lists
            for _, attr_value in vars(obj).items():
                if isinstance(attr_value, list):
                    for item in attr_value:
                        if hasattr(item, "elset"):
                            elset = getattr(item, "elset")
                            if elset is not None:
                                used.add(elset.index)

        return used

    # ==========================================================
    # NAME LOOKUP UTILITIES
    # ==========================================================
    def _ensure_name_index(self):
        """
        Build a name → object lookup table.
        Only objects with a .name attribute are included.
        """
        self._name_index = {}

        for obj in self.objects:
            if hasattr(obj, "name"):
                self._name_index[obj.name] = obj

    def get_by_name(self, name: str):
        """
        Soft name lookup.
        Returns the object whose .name equals `name`, or None if not found.

        Raises:
            AttributeError if objects do not have .name at all.
        """

        if not self.objects:
            return None

        first = self.objects[0]
        if not hasattr(first, "name"):
            raise AttributeError(
                f"{type(self).__name__} objects do not define a 'name' attribute"
            )

        # rebuild table each time (safe & fresh)
        for obj in self.objects:
            if hasattr(obj, "name") and obj.name == name:
                return obj

        return None  # <── IMPORTANT

class CollectionParser(ABC, Generic[M, T, C]):

    @classmethod
    @abstractmethod
    def get_collection(cls) -> Type[C]:
        pass

    @classmethod
    @abstractmethod
    def parse_line(cls, lines: List[str], **kwargs) -> T:
        pass

    @classmethod
    def from_model(cls, model: M, **kwargs) -> C:
        collection_cls = cls.get_collection()
        block = model.blocks.get(collection_cls.header)
        parsed_items: list[T] = []

        n = getattr(cls, "LINES_PER_ITEM", 1)
        lines = block.body

        for i in range(0, len(lines), n):
            item_lines = lines[i:i + n]
            try:
                parsed_item = cls.parse_line(item_lines, **kwargs)
            except Exception:
                import traceback; traceback.print_exc()
                parsed_item = None

            if parsed_item:
                parsed_items.append(parsed_item)

        return collection_cls(parsed_items)
    
class ObjectCollectionAdapter(ABC, Generic[M, T, C]):
  
    @classmethod
    @abstractmethod
    def update_var(cls, collection: C, model: M) -> M:
        return model

    @classmethod
    @abstractmethod
    def format_line(cls, obj: T) -> str:
        pass
    
    @classmethod
    def to_string(cls, collection: C) -> str:
        lines = [f'*{collection.header}*']
        for obj in collection.objects:
            lines.append(cls.format_line(obj))
        return "\n".join(lines)
    
    @classmethod
    def to_block(cls, collection: C) -> str:
        header = collection.header
        lines = []
        for obj in collection.objects:
            lines.append(cls.format_line(obj))
        return BlockAdapter.from_lines(header= header, lines= lines)
    
    @classmethod
    def to_model(cls, collection: C, model: M) -> M:
        header = collection.header
        lines = []

        for obj in collection.objects:
            lines.append(cls.format_line(obj))

        block = BlockAdapter.from_lines(header= header, lines= lines)
        model.blocks[header] = block
        model = cls.update_var(collection, model)
        return model
    
    @classmethod
    def _norm_float(cls, value) -> Union[int, float]:
        value = float(value)
        return int(value) if value.is_integer() else value
    
    @classmethod
    def _norm_float_sci(cls, value: float):
        if value.is_integer():
            return int(value)

        if abs(value) < 1e-3 or abs(value) >= 1e4:
            s = f"{value:.15E}"
            base, exp = s.split("E")
            base = base.rstrip("0").rstrip(".")
            return f"{base}E{int(exp):+04d}"

        return value
    
    # -------------------------------------------------------------
    # EXCEL EXPORT
    # -------------------------------------------------------------
    
    @classmethod
    def _to_excel_safe(cls, value):
        if value is None:
            return ""
        if isinstance(value, float):
            return cls._norm_float_sci(value)
        return value

    @classmethod
    def _flatten(cls, obj, prefix="", visited=None, depth=0, *, root=True):
        if visited is None:
            visited = set()
        if depth > 20:
            return {prefix[:-1]: "<DepthLimit>"}

        oid = id(obj)
        if oid in visited:
            return {prefix[:-1]: "<CircularRef>"}
        visited.add(oid)

        # ------------------------------------------------------------
        # 1) Primitive
        # ------------------------------------------------------------
        if isinstance(obj, (str, int, float, bool)) or obj is None:
            return {prefix[:-1]: obj}

        # ------------------------------------------------------------
        # 2) COLLAPSE nested Object-subclasses to index
        # ------------------------------------------------------------
        # NOTE: this must be BEFORE any asdict/vars()
        if is_dataclass(obj) and isinstance(obj, Object) and not root:
            return {prefix[:-1]: obj.index}

        # ------------------------------------------------------------
        # 3) Expand dataclass normally (but only here)
        # ------------------------------------------------------------
        if is_dataclass(obj):
            obj = obj.__dict__     # <-- NOT asdict() !!!
        elif hasattr(obj, "__dict__") and not isinstance(obj, type):
            obj = obj.__dict__
        elif isinstance(obj, dict):
            obj = obj
        else:
            return {prefix[:-1]: obj}

        # ------------------------------------------------------------
        # 4) Recurse into fields
        # ------------------------------------------------------------
        flat = {}
        for key, value in obj.items():
            name = f"{prefix}{key}"

            if (
                is_dataclass(value)
                or hasattr(value, "__dict__")
                or isinstance(value, dict)
            ):
                flat.update(
                    cls._flatten(
                        value,
                        prefix=f"{name}.",
                        visited=visited,
                        depth=depth+1,
                        root=False,
                    )
                )
            else:
                flat[name] = value

        return flat


    @classmethod
    def _prune_empty_nested(cls, flat: dict) -> dict:
        """
        Remove nested group columns (e.g. node.index, node.x, node.y, node.z)
        if the group has no meaningful data except index.
        """
        groups = {}

        for key, val in flat.items():
            if "." not in key:
                continue
            root, child = key.split(".", 1)
            groups.setdefault(root, {})[child] = val

        to_delete = []

        for root, children in groups.items():
            idx_key = f"{root}.index"
            if idx_key not in flat:
                continue

            payload = {k: v for k, v in children.items() if k != "index"}

            # define empty nested: all None/""/0
            empty = all(v in (None, "", 0) for v in payload.values())

            if empty:
                for k in [f"{root}.{c}" for c in children.keys()]:
                    to_delete.append(k)

        for k in to_delete:
            flat.pop(k, None)

        return flat
    
    @classmethod
    def _infer_headers_from_type(cls, typ) -> List[str]:
        if not is_dataclass(typ):
            raise TypeError(f"Expected dataclass type, got {typ}")

        headers = []

        for f in typ.__dataclass_fields__.values():
            t = f.type

            if is_dataclass(t):
                sub = cls._infer_headers_from_type(t)
                headers.extend([f"{f.name}.{x}" for x in sub])
            else:
                headers.append(f.name)

        return headers
    
    @staticmethod
    def _normalize_excel_value(v):
        """
        Excel cannot store tuples, lists, dicts, or custom objects directly.
        Convert tuples into CSV strings (e.g. (1,2,3) -> '1,2,3').
        Leave other primitive types untouched.
        """
        if isinstance(v, tuple):
            return ",".join(str(x) for x in v)
        return v

    @classmethod
    def export_to_excel(
        cls,
        collections: List[Tuple[str, C]],
        folder_path: str,
        excel_name: str,
    ):
        if not collections:
            raise ValueError("No collections provided")

        os.makedirs(folder_path, exist_ok=True)
        filepath = os.path.join(folder_path, f"{excel_name}.xlsx")

        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, collection in collections:
            print(collection.header)
            ws = wb.create_sheet(title=sheet_name)

            objs = collection.objects

            # ================================================
            # Determine headers
            # ================================================
            if objs:
                first_flat = cls._flatten(objs[0])
                headers = list(first_flat.keys())
            else:
                obj_type = getattr(collection, "item_type", None)
                if obj_type is None:
                    raise ValueError(f"Collection '{sheet_name}' has no item_type")
                headers = cls._infer_headers_from_type(obj_type)

            # Write headers always
            ws.append(headers)

            # ================================================
            # Rows for non-empty collections
            # ================================================
            for obj in objs:
                flat = cls._flatten(obj)
                row = [
                    cls._normalize_excel_value(flat.get(h, ""))
                    for h in headers
                ]
                ws.append(row)

        wb.save(filepath)
        print(f"✅ Exported {len(collections)} collections → {filepath}")


    # -------------------------------------------------------------
    # IMPORT EXCEL
    # -------------------------------------------------------------

    # @classmethod
    # def _resolve_references(cls, collections: dict):
    #     """
    #     Replace placeholder nested objects with real instances,
    #     matched by (type, index).
    #     """

    #     # Build global lookup
    #     lookup = {}

    #     for coll in collections.values():
    #         if not hasattr(coll, "objects"):
    #             continue

    #         for obj in coll.objects:
    #             if isinstance(obj, Object):
    #                 lookup[(type(obj), obj.index)] = obj

    #     # Resolve nested references
    #     for coll in collections.values():
    #         for obj in coll.objects:
    #             for f in fields(type(obj)):
    #                 val = getattr(obj, f.name)

    #                 # is it a placeholder nested object?
    #                 if isinstance(val, Object) and val.index is not None:
    #                     key = (type(val), val.index)

    #                     if key in lookup:
    #                         setattr(obj, f.name, lookup[key])
    #                     else:
    #                         print(f"⚠ Missing reference: {type(val).__name__}[{val.index}]")


    @classmethod
    def _resolve_elsets(cls, collections: dict, elsets: Collection):
        """
        Replace any `.elset` placeholder with the real Elset.
        Report only once at the end whether resolution succeeded or failed.
        """

        elset_map = {e.index: e for e in elsets.objects}

        total = 0
        resolved = 0
        failed = 0

        for coll_name, coll in collections.items():
            if not hasattr(coll, "objects"):
                continue

            for obj in coll.objects:

                # only objects that actually use an elset
                if not hasattr(obj, "elset"):
                    continue

                placeholder = obj.elset
                if placeholder is None:
                    continue

                total += 1

                # extract index
                try:
                    idx = placeholder.index
                except AttributeError:
                    idx = placeholder  # raw int fallback

                # attempt resolution
                real = elset_map.get(idx)
                if real is None:
                    failed += 1
                    continue

                obj.elset = real
                resolved += 1

        # --------------------------------------
        # FINAL SUMMARY NOTICE (the important part)
        # --------------------------------------
        if total == 0:
            print("ℹ No elset references to resolve.")
            return

        if failed == 0:
            print(f"✓ All elsets resolved successfully ({resolved}/{total}).")
        else:
            print(
                f"⚠ Elset resolution incomplete: "
                f"{resolved}/{total} resolved, {failed} missing."
            )


    @classmethod
    def _resolve_references(cls, collections: dict):
        """
        Replace placeholder nested objects with real instances,
        matched by (type, index).
        """

        # Build global lookup
        lookup = {}

        for coll_name, coll in collections.items():
            if not hasattr(coll, "objects"):
                continue

            for obj in coll.objects:
                if isinstance(obj, Object):
                    lookup[(type(obj), obj.index)] = obj

        # Resolve nested references
        for coll_name, coll in collections.items():
            if not hasattr(coll, "objects"):
                continue

            for obj in coll.objects:
                for f in fields(type(obj)):
                    val = getattr(obj, f.name)

                    # is it a placeholder nested object?
                    if isinstance(val, Object) and val.index is not None:
                        key = (type(val), val.index)

                        if key in lookup:
                            setattr(obj, f.name, lookup[key])
                        else:
                            parent_type = type(obj).__name__
                            parent_index = getattr(obj, "index", "?")
                            field_name = f.name
                            missing_type = type(val).__name__
                            missing_index = val.index

                            print(
                                f"⚠ Missing reference in {coll_name}: "
                                f"{parent_type}[{parent_index}].{field_name} → "
                                f"{missing_type}[{missing_index}] not found"
                            )

    @classmethod
    def _from_excel_row(cls, data: dict, dataclass_type):
        kwargs = {}

        for field in fields(dataclass_type):
            name = field.name

            if name not in data:
                continue

            value = data[name]

            # Nested Object field
            if issubclass(field.type, Object):
                if value in ("", None):
                    kwargs[name] = None
                else:
                    sub = field.type.__new__(field.type)
                    sub.index = int(value)

                    # all other attributes become None
                    for f in fields(field.type):
                        if f.name != "index":
                            setattr(sub, f.name, None)

                    kwargs[name] = sub

            # Primitive field
            else:
                if field.type is int:
                    kwargs[name] = int(value)
                elif field.type is float:
                    kwargs[name] = float(value)
                else:
                    kwargs[name] = value

        return dataclass_type(**kwargs)


    @classmethod
    def _import_sheet(cls, sheet, dataclass_type):
        rows = list(sheet.values)
        if not rows:
            return []

        headers = rows[0]
        objects = []

        for row in rows[1:]:
            if row is None or all(v is None for v in row):
                continue

            data = {headers[i]: row[i] for i in range(len(headers))}
            obj = cls._from_excel_row(data, dataclass_type)
            objects.append(obj)

        return objects


    @classmethod
    def from_excel(cls, folder_path: str, excel_name: str, mapping: dict, elsets:Collection = None):
        """
        mapping = {
            "Nodes": Nodes,      # collection class
            "Offsets": Offsets,
        }
        """
        from openpyxl import load_workbook
        import os

        filepath = os.path.join(folder_path, f"{excel_name}.xlsx")
        if not os.path.exists(filepath):
            raise FileNotFoundError(filepath)

        wb = load_workbook(filepath, data_only=True)
        collections = {}

        # ---------------------------------------------------------
        # 1) Import each sheet → build collections
        # ---------------------------------------------------------
        for sheet_name, collection_cls in mapping.items():

            if sheet_name not in wb.sheetnames:
                print(f"⚠ Sheet '{sheet_name}' not found in '{excel_name}.xlsx'")
                continue

            sheet = wb[sheet_name]
            dataclass_type = collection_cls.item_type

            objects = cls._import_sheet(sheet, dataclass_type)
            collections[sheet_name] = collection_cls(objects=objects)

        # ---------------------------------------------------------
        # 2) Auto-resolve references across all collections
        # ---------------------------------------------------------
        cls._resolve_references(collections)

        # 3. bind elsets if provided
        if elsets is not None:
            cls._resolve_elsets(collections, elsets)

        return collections

class ObjectCollectionQuery(ABC, Generic[T, C]):
    TOL = 1e-6

    @staticmethod
    @abstractmethod
    def get_by_indices(collection: C, indices: List[int]) -> C:
        raise NotImplementedError

    @staticmethod
    @abstractmethod
    def get_by_offset(collection: C, dx: float, dy: float, dz: float,
                      origin: Optional[T] = None) -> Optional[T]:
        raise NotImplementedError

    @staticmethod
    @abstractmethod
    def select_by_polygon(collection: C, boundary_indices: List[int]) -> C:
        raise NotImplementedError

class ObjectCollectionEngine(ABC, Generic[T, C]):

    @staticmethod
    @abstractmethod
    def replicate(base_collection: C,
                  collection_to_copy: C,
                  *args, **kwargs) -> C:
        """Replicate selected_objects within base_collection under given params."""
        pass

    @staticmethod
    @abstractmethod
    def extend(base_collection: C,
               target_index: int,
               *args, **kwargs) -> C:
        """Extend base_collection with new objects up to target_index."""
        pass

class CollectionComparer(ABC, Generic[M, T, C]):

    def __init__(self, existing, imported):
        self.existing  = existing
        self.imported  = imported

    # ------------------------------------------------------------------
    # Default sort: by name (case-insensitive) or fallback to +inf
    # ------------------------------------------------------------------
    @staticmethod
    def get_sort_key(obj):
        if hasattr(obj, "name"):
            return obj.name.lower()
        return float("inf")

    # ------------------------------------------------------------------
    # MAIN MERGE PIPELINE
    # ------------------------------------------------------------------
    def merge_and_reorder(
        self,
        key: Callable[[T], Any] = None,
        unique_attr: str = "name",
        remove_missing: bool = False,
        used_elsets: Optional[Set[int]] = None,
    ) -> Tuple[C, Dict[int, int], bool]:

        key = key or self.get_sort_key
        used_elsets = used_elsets or set()

        existing_lookup = {
            getattr(o, unique_attr): o for o in self.existing.objects
        }
        imported_lookup = {
            getattr(o, unique_attr): o for o in self.imported.objects
        }

        # Store original indices before mutation
        old_index_by_name = {
            name: obj.index for name, obj in existing_lookup.items()
        }

        merged: list[T] = []

        # ================================================================
        # 1) EXISTING ITEMS — REMOVE_MISSING PROTOCOL
        # ================================================================
        for name, old_obj in existing_lookup.items():

            if name in imported_lookup:
                # A → imported overwrites existing
                imp_obj = imported_lookup[name]
                self._copy_attributes_overwriting(old_obj, imp_obj)
                merged.append(old_obj)
                continue

            # B → missing in imported

            if not remove_missing:
                # remove_missing=False → KEEP ALL existing
                merged.append(old_obj)
                continue

            # remove_missing=True → remove UNLESS protected by used_elsets
            if old_obj.index in used_elsets:
                merged.append(old_obj)
                continue

            # remove_missing=True + unused → drop
            pass

        # ================================================================
        # 2) NEW IMPORTED ITEMS (names not in existing)
        # ================================================================
        for name, imp_obj in imported_lookup.items():
            if name not in existing_lookup:
                merged.append(imp_obj)

        # ================================================================
        # 3) SORT MERGED RESULT
        # ================================================================
        merged_sorted = sorted(merged, key=key)

        # ================================================================
        # 4) BUILD NEW INDEX MAP (name → new index)
        # ================================================================
        name_to_new_index = {
            getattr(obj, unique_attr): idx
            for idx, obj in enumerate(merged_sorted, start=1)
        }

        # ================================================================
        # 5) REORDER MAP (old idx → new idx)
        # Only for items that still exist after merge
        # ================================================================
        reorder_map: dict[int, int] = {
            old_idx: name_to_new_index[name]
            for name, old_idx in old_index_by_name.items()
            if name in name_to_new_index
        }

        # ================================================================
        # 6) REASSIGN NEW INDICES (mutate objects)
        # ================================================================
        for idx, obj in enumerate(merged_sorted, start=1):
            obj.index = idx

        # ================================================================
        # 7) BUILD NEW COLLECTION
        # ================================================================
        collection_cls = type(self.existing)
        merged_collection = collection_cls(merged_sorted)

        # ================================================================
        # 8) REMOVAL FLAG
        # Detect ANY removed existing item accurately
        # ================================================================
        removed_any = any(
            name not in name_to_new_index 
            for name in existing_lookup
        )

        return merged_collection, reorder_map, removed_any

    # ------------------------------------------------------------------
    # Attribute overwrite: imported → existing (never touch index)
    # ------------------------------------------------------------------
    @staticmethod
    def _copy_attributes_overwriting(existing_obj, imported_obj):
        for attr, value in imported_obj.__dict__.items():
            if attr == "index":
                continue
            setattr(existing_obj, attr, value)