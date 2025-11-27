from dataclasses import dataclass, field
from typing import List, Generic, TypeVar
from object._object_abstract import Object
from layout._layout_abstract import LayoutBase

I = TypeVar("I", bound="Object") 
L = TypeVar("L", bound="LayoutBase")

@dataclass
class CompactLayoutBase(Generic[I]):
    index: int
    items: List[I] = field(default_factory=list)

    def add(self, item: I):
        self.items.append(item)

    def __len__(self): return len(self.items)
    def __iter__(self): return iter(self.items)

    def as_parent_row(self) -> dict:
        """Return the parent (layout) as a flat dict."""
        return {"index": self.index}

    def as_child_rows(self) -> List[dict]:
        """Return each item as a dict."""
        rows = []
        for it in self.items:
            if hasattr(it, "__dict__"):
                rows.append(it.__dict__)
            else:
                rows.append({"value": it})
        return rows

import os
from typing import List
from openpyxl import Workbook

class CompactLayoutsBase(Generic[L]):
    def __init__(self, layouts=None):
        self.layouts: List[L] = layouts or []
        self._by_index = {lay.index: lay for lay in self.layouts}

    def add(self, layout: L):
        self.layouts.append(layout)
        self._by_index[layout.index] = layout

    def get(self, index: int) -> L:
        return self._by_index[index]

    def __iter__(self):
        return iter(self.layouts)

    # ---------------------------------------------------------
    #  INTERNALIZED EXCEL SHEET SERIALIZER
    # ---------------------------------------------------------
    def to_excel_sheets(self):
        for layout in self.layouts:
            parent = layout.as_parent_row()
            children = layout.as_child_rows()

            rows = []
            rows.append(list(parent.keys()))
            rows.append(list(parent.values()))
            rows.append([])

            if children:
                rows.append(["items:"])
                headers = list(children[0].keys())
                rows.append(headers)
                for c in children:
                    rows.append([c.get(h) for h in headers])

            yield (f"Layout_{layout.index}", rows)

    # ---------------------------------------------------------
    #  INTERNALIZED EXCEL EXPORT
    # ---------------------------------------------------------
    def export_to_excel(self, folder, excel_name):
        os.makedirs(folder, exist_ok=True)
        filepath = os.path.join(folder, excel_name + ".xlsx")

        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, rows in self.to_excel_sheets():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)

        wb.save(filepath)
        print(f"✓ Exported → {filepath}")

    @classmethod
    def from_excel(
        cls,
        folder: str,
        excel_name: str,
        layout_cls,
        item_cls
    ):
        """
        Load CompactLayoutsBase from an Excel workbook.

        Args:
            folder: folder containing the .xlsx
            excel_name: filename without extension
            layout_cls: class for each layout (e.g. CompactBeamLayout)
            item_cls: dataclass for child items (e.g. BeamCompact)
        """

        filepath = os.path.join(folder, f"{excel_name}.xlsx")

        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)

        layouts = []

        for sheet in wb.worksheets:

            # Convert rows to clean Python lists
            rows = [[c for c in r] for r in sheet.iter_rows(values_only=True)]
            if not rows:
                continue

            # -------------------------------------------------
            # 1) PARENT SECTION
            # -------------------------------------------------
            parent_headers = rows[0]
            parent_values  = rows[1]

            parent = dict(zip(parent_headers, parent_values))
            layout_index = int(parent["index"])

            # -------------------------------------------------
            # 2) FIND CHILDREN BLOCK ("items:")
            # -------------------------------------------------
            child_start = None
            for i, row in enumerate(rows):
                if row and row[0] == "items:":
                    child_start = i
                    break

            items = []
            if child_start is not None:
                headers = rows[child_start + 1]

                for r in rows[child_start + 2:]:
                    if all(v is None for v in r):
                        break
                    row_dict = dict(zip(headers, r))

                    # force layout field to match parent layout
                    if "layout" in row_dict:
                        row_dict["layout"] = layout_index

                    items.append(item_cls(**row_dict))

            # -------------------------------------------------
            # 3) BUILD LAYOUT OBJECT
            # -------------------------------------------------
            layout_obj = layout_cls(index=layout_index, items=items)
            layouts.append(layout_obj)

        return cls(layouts=layouts)

