import os
from dataclasses import fields, is_dataclass
from enum import Enum
from openpyxl import load_workbook
from typing import Any, Dict, List, Tuple, Type


def _set_nested(data: dict, path: str, value: Any):
    """Rebuild nested structure from dot-separated keys."""
    keys = path.split(".")
    cur = data
    for k in keys[:-1]:
        cur = cur.setdefault(k, {})
    cur[keys[-1]] = value


def _coerce_value(value: str) -> Any:
    """Try to infer numeric/bool/empty types."""
    if value == "" or value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, str):
        val = value.strip()
        # bool
        if val.lower() in ("true", "false"):
            return val.lower() == "true"
        # int
        try:
            if "." not in val:
                return int(val)
        except ValueError:
            pass
        # float
        try:
            return float(val)
        except ValueError:
            pass
    return value


def _expand_nested(flat_dict: Dict[str, Any]) -> Dict[str, Any]:
    """Expand flattened dict into nested dict structure."""
    nested = {}
    for key, val in flat_dict.items():
        _set_nested(nested, key, val)
    return nested


def _dict_to_dataclass(cls: Type, data: Dict[str, Any]):
    """Instantiate dataclass (recursively) from dict."""
    if not is_dataclass(cls):
        raise TypeError(f"{cls} is not a dataclass")

    kwargs = {}
    for f in fields(cls):
        val = data.get(f.name)
        if val is None:
            kwargs[f.name] = None
            continue

        # nested dataclass
        if is_dataclass(f.type) and isinstance(val, dict):
            kwargs[f.name] = _dict_to_dataclass(f.type, val)
        # Enum
        elif isinstance(f.type, type) and issubclass(f.type, Enum):
            try:
                kwargs[f.name] = f.type[val]
            except KeyError:
                kwargs[f.name] = None
        else:
            kwargs[f.name] = _coerce_value(val)
    return cls(**kwargs)


def import_multiple_collections_from_excel(
    filepath: str,
    sheet_to_class: Dict[str, Type],
) -> Dict[str, List[object]]:
    """
    Read Excel file and rebuild collections.

    Args:
        filepath: path to Excel file
        sheet_to_class: mapping {sheet_name: dataclass_type}
    Returns:
        dict: {sheet_name: [objects]}
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(filepath)

    wb = load_workbook(filepath, data_only=True)
    results = {}

    for sheet_name, cls in sheet_to_class.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = rows[0]
        objects = []

        for row in rows[1:]:
            flat = {h: _coerce_value(v) for h, v in zip(headers, row)}
            nested = _expand_nested(flat)
            obj = _dict_to_dataclass(cls, nested)
            objects.append(obj)

        results[sheet_name] = objects

    return results

def add_prefix_dict_keys(data: dict, prefix: str) -> dict:
    """Return a new dict with all keys prefixed."""
    return {f"{prefix}{k}": v for k, v in data.items()}

# ================================================================
# Convert imported lists → proper SANSPRO collections
# ================================================================

from collection._collection_abstract import Collection
from typing import TypeVar, Type, List

T = TypeVar("T")

def import_collection_from_excel(filepath: str,
                                 sheet_name: str,
                                 cls: Type[T]) -> List[T]:
    """
    Read one sheet and convert each row into dataclass cls.
    """
    wb = load_workbook(filepath, data_only=True)

    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    header = rows[0]
    objects = []

    for row in rows[1:]:
        flat = {h: _coerce_value(v) for h, v in zip(header, row)}
        nested = _expand_nested(flat)
        obj = _dict_to_dataclass(cls, nested)
        objects.append(obj)

    return objects

def import_multiple_collections_from_excel(
        filepath: str,
        sheet_map: Dict[str, Type]
    ) -> Dict[str, List[object]]:
    
    results = {}
    for sheet, cls in sheet_map.items():
        results[sheet] = import_collection_from_excel(filepath, sheet, cls)
    return results

def build_collections_from_import(data: Dict[str, List[object]]) -> Dict[str, Collection]:
    out = {}
    for name, objects in data.items():
        if not objects:
            out[name] = None
            continue

        # infer collection type: Node → Nodes, Bar → Bars, etc.
        obj_type = type(objects[0])
        coll_type_name = obj_type.__name__ + "s"
        coll_type = globals().get(coll_type_name)

        if not coll_type:
            raise RuntimeError(f"No collection class found for {obj_type}")

        out[name] = coll_type(objects)

    return out

def import_sanspro_collections(
    filepath: str,
    sheet_map: Dict[str, Tuple[type, type]]
):
    raw = {}

    # parse Excel → object lists
    for sheet_name, (obj_cls, _) in sheet_map.items():
        raw[sheet_name] = import_collection_from_excel(filepath, sheet_name, obj_cls)

    # wrap result into collection classes
    collections = {}

    for sheet_name, (_, coll_cls) in sheet_map.items():
        objects = raw[sheet_name]
        collections[sheet_name] = coll_cls(objects)  # <── direct and explicit

    return collections
