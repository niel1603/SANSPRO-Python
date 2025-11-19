# from openpyxl import Workbook
# import os
# from typing import List, Tuple

# def export_multiple_collections_to_excel(collections: List[Tuple[str, List[object]]],
#                                          folder_path: str,
#                                          excel_name: str):
#     if not collections:
#         raise ValueError("No collections provided")

#     os.makedirs(folder_path, exist_ok=True)
#     filepath = os.path.join(folder_path, f"{excel_name}.xlsx")

#     wb = Workbook()
#     wb.remove(wb.active)  # remove default sheet

#     for sheet_name, objects in collections:
#         if not objects:
#             continue

#         ws = wb.create_sheet(title=sheet_name)
#         headers = list(vars(objects[0]).keys())
#         ws.append(headers)

#         for obj in objects:
#             row = []
#             for attr_name in headers:
#                 attr = getattr(obj, attr_name)
#                 if hasattr(attr, "index"):  # If it's an object with .index
#                     row.append(getattr(attr, "index"))
#                 else:
#                     row.append(attr)
#             ws.append(row)

#     wb.save(filepath)

import os
from dataclasses import is_dataclass, asdict
from enum import Enum
from typing import List, Tuple
from openpyxl import Workbook


def _to_excel_safe(value):
    """Convert value to something Excel can store."""
    if value is None:
        return ""
    if isinstance(value, (int, float, str, bool)):
        return value
    if isinstance(value, Enum):
        return value.name
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return "; ".join(f"{k}={v}" for k, v in value.items())
    return str(value)


def _flatten(obj, prefix="", visited=None, depth=0):
    """Flatten dataclass or object safely into {col: value}."""
    if visited is None:
        visited = set()
    if depth > 10:
        # prevent runaway recursion
        return {prefix[:-1]: f"<DepthLimit@{depth}>"}

    obj_id = id(obj)
    if obj_id in visited:
        return {prefix[:-1]: "<CircularRef>"}
    visited.add(obj_id)

    data = {}

    # Handle enums and basic types early
    if isinstance(obj, (Enum, str, int, float, bool)) or obj is None:
        return {prefix[:-1]: _to_excel_safe(obj)}

    # Dataclasses
    if is_dataclass(obj):
        obj = asdict(obj)
    elif hasattr(obj, "__dict__") and not isinstance(obj, type):
        obj = vars(obj)
    elif isinstance(obj, dict):
        obj = obj
    else:
        return {prefix[:-1]: _to_excel_safe(obj)}

    for key, value in obj.items():
        name = f"{prefix}{key}"
        if is_dataclass(value) or (
            hasattr(value, "__dict__")
            and not isinstance(value, (Enum, type, str, bytes))
        ) or isinstance(value, dict):
            data.update(_flatten(value, prefix=f"{name}.", visited=visited, depth=depth + 1))
        else:
            data[name] = _to_excel_safe(value)

    return data


def export_multiple_collections_to_excel(
    collections: List[Tuple[str, List[object]]],
    folder_path: str,
    excel_name: str,
):
    if not collections:
        raise ValueError("No collections provided")

    os.makedirs(folder_path, exist_ok=True)
    filepath = os.path.join(folder_path, f"{excel_name}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, objects in collections:
        if not objects:
            continue

        ws = wb.create_sheet(title=sheet_name)
        first_flat = _flatten(objects[0])
        headers = list(first_flat.keys())
        ws.append(headers)

        for obj in objects:
            flat = _flatten(obj)
            row = [flat.get(h, "") for h in headers]
            ws.append(row)

    wb.save(filepath)
    print(f"✅ Exported {len(collections)} collections → {filepath}")

def strip_prefix_dict_keys(data: dict, prefix: str) -> dict:
    """Return a new dict with prefix removed from keys (if present)."""
    plen = len(prefix)
    return {
        (k[plen:] if k.startswith(prefix) else k): v
        for k, v in data.items()
    }